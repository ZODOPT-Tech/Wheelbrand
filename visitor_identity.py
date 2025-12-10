import streamlit as st
import mysql.connector
import boto3
import json
from datetime import datetime
import base64
import io
from PIL import Image as PILImage
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as XLImage


# =============================
# AWS Configuration
# =============================
AWS_REGION = "ap-south-1"
AWS_BUCKET = "zodoptvisiorsmanagement"
EXCEL_FILE = "visitorsphoto.xlsx"
AWS_SECRET_ARN = "arn:aws:secretsmanager:ap-south-1:034362058776:secret:Wheelbrand-zM6npS"


# =============================
# DB Credentials from AWS
# =============================
@st.cache_resource
def get_db_credentials():
    client = boto3.client("secretsmanager", region_name=AWS_REGION)
    sec = client.get_secret_value(SecretId=AWS_SECRET_ARN)
    return json.loads(sec["SecretString"])


@st.cache_resource
def get_db_conn():
    c = get_db_credentials()
    return mysql.connector.connect(
        host=c["DB_HOST"],
        user=c["DB_USER"],
        password=c["DB_PASSWORD"],
        database=c["DB_NAME"],
        autocommit=True,
    )


# =============================
# APPROVE Visitor Once Pass Generated
# =============================
def approve_visitor(visitor_id):
    conn = get_db_conn()
    cur = conn.cursor()
    cur.execute(
        "UPDATE visitors SET status='approved' WHERE visitor_id=%s",
        (visitor_id,)
    )
    cur.close()


# =============================
# Get Visitor Details from DB
# =============================
def get_visitor_data(visitor_id: int):
    conn = get_db_conn()
    cur = conn.cursor(dictionary=True)

    cur.execute("""
        SELECT 
            visitor_id,
            full_name,
            from_company,
            person_to_meet,
            visit_type,
            registration_timestamp
        FROM visitors
        WHERE visitor_id = %s;
    """, (visitor_id,))

    data = cur.fetchone()
    cur.close()
    return data


# =============================
# Store Photo in S3 Excel File (Safe Version)
# =============================
def write_photo_to_excel(visitor, photo_bytes):
    s3 = boto3.client("s3")

    # Step 1: Try to load existing Excel
    try:
        obj = s3.get_object(Bucket=AWS_BUCKET, Key=EXCEL_FILE)
        xl_data = obj["Body"].read()
        wb = load_workbook(io.BytesIO(xl_data))
        ws = wb.active

        # if sheet empty -> add headers
        if ws.max_row == 1 and not ws["A1"].value:
            ws["A1"] = "Visitor Name"
            ws["B1"] = "Company"
            ws["C1"] = "Photo"
            ws["D1"] = "Timestamp"

    except Exception:
        # Step 2: Create new workbook if missing/invalid
        wb = Workbook()
        ws = wb.active
        ws.title = "Visitors"
        ws["A1"] = "Visitor Name"
        ws["B1"] = "Company"
        ws["C1"] = "Photo"
        ws["D1"] = "Timestamp"

    # Step 3: Append row
    next_row = ws.max_row + 1
    ws[f"A{next_row}"] = visitor["full_name"]
    ws[f"B{next_row}"] = visitor["from_company"]
    ws[f"D{next_row}"] = datetime.now().strftime("%Y-%m-%d %H:%M")

    # Step 4: Add image in cell C
    img = PILImage.open(io.BytesIO(photo_bytes))
    img.thumbnail((120, 120))
    xl_img = XLImage(img)
    xl_img.anchor = f"C{next_row}"
    ws.add_image(xl_img)

    # Step 5: Save workbook to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Step 6: Upload back to S3
    s3.put_object(
        Bucket=AWS_BUCKET,
        Key=EXCEL_FILE,
        Body=output.getvalue(),
        ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# =============================
# Header (Like visitor_details)
# =============================
def render_header(visitor):
    st.markdown("""
        <style>
            .id-header {
                background: linear-gradient(90deg, #5036FF, #9C2CFF);
                padding: 24px;
                border-radius: 14px;
                color: white;
                font-size: 26px;
                font-weight: 700;
                margin-bottom: 18px;
            }
            .id-sub {
                font-size: 15px;
                opacity: 0.92;
                margin-top: 6px;
            }
            .profile-box {
                background:white;
                padding:16px;
                border-radius:10px;
                box-shadow:0px 3px 8px rgba(0,0,0,0.1);
                margin-bottom:12px;
            }
        </style>
    """, unsafe_allow_html=True)

    st.markdown(f"""
        <div class="id-header">
            Visitor Identity Capture
            <div class="id-sub">Take photo & generate secure visitor pass</div>
        </div>
    """, unsafe_allow_html=True)

    st.markdown(f"""
        <div class="profile-box">
            <b>Name:</b> {visitor['full_name']}<br>
            <b>Company:</b> {visitor['from_company']}<br>
            <b>Meeting:</b> {visitor['person_to_meet']}<br>
        </div>
    """, unsafe_allow_html=True)


# =============================
# HTML Visitor Pass
# =============================
def show_pass(visitor, photo_bytes):
    b64 = base64.b64encode(photo_bytes).decode()

    st.markdown(f"""
        <div style="width:420px;
                    background:white;
                    box-shadow:0 4px 12px rgba(0,0,0,0.14);
                    border-radius:12px;
                    padding:20px;
                    margin-top:20px;">
            <h2 style="text-align:center;color:#5036FF;">VISITOR PASS</h2>
            
            <div style="text-align:center;margin-bottom:10px;">
                <img src="data:image/jpeg;base64,{b64}"
                     style="width:140px;height:140px;
                            border-radius:10px;
                            border:2px solid #5036FF;">
            </div>

            <p><b>Name:</b> {visitor['full_name']}</p>
            <p><b>From:</b> {visitor['from_company']}</p>
            <p><b>To Meet:</b> {visitor['person_to_meet']}</p>
            <p><b>Visitor ID:</b> #{visitor['visitor_id']}</p>
            <p><b>Date:</b> {datetime.now().strftime("%d-%m-%Y %H:%M")}</p>
        </div>
    """, unsafe_allow_html=True)


# =============================
# MAIN ENTRY EXPORT
# =============================
def render_identity_page():

    # Auth Check
    if not st.session_state.get("admin_logged_in", False):
        st.session_state["current_page"] = "visitor_login"
        st.rerun()

    # Ensure visitor selected
    if "current_visitor_id" not in st.session_state:
        st.session_state["current_page"] = "visitor_details"
        st.rerun()

    visitor_id = st.session_state["current_visitor_id"]
    visitor = get_visitor_data(visitor_id)

    # Header
    render_header(visitor)

    # Camera
    photo = st.camera_input("Capture Photo")

    if st.button("Save & Generate Pass"):
        if not photo:
            st.error("Please capture the photo.")
            return

        photo_bytes = photo.getvalue()

        with st.spinner("Saving identity and generating pass..."):
            write_photo_to_excel(visitor, photo_bytes)
            approve_visitor(visitor_id)

        st.success("Visitor registered successfully!")
        show_pass(visitor, photo_bytes)

        col1, col2, col3 = st.columns(3)

        with col1:
            if st.button("New Visitor"):
                st.session_state.pop("current_visitor_id", None)
                st.session_state["current_page"] = "visitor_details"
                st.rerun()

        with col2:
            if st.button("Dashboard"):
                st.session_state["current_page"] = "visitor_dashboard"
                st.rerun()

        with col3:
            if st.button("Logout"):
                st.session_state.clear()
                st.session_state["current_page"] = "visitor_login"
                st.rerun()

    if st.button("Back"):
        st.session_state["current_page"] = "visitor_dashboard"
        st.rerun()
